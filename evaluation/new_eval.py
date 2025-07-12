import argparse
import os
import json
import time
import re
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from langchain_openai import AzureOpenAI,OpenAI,AzureChatOpenAI,ChatOpenAI

SYSTEM_PROMPT = """你是一個分析師，需要分析給定的任務是否成功並提取規則以幫助未來的代理人完成任務。

你將根據以下內容進行分析：

1. 網頁任務
2. 操作紀錄
3. 操作紀錄截圖
4. 任務答案

# Steps

1. **評估網頁任務指令**：查看指令是否包含多個任務，確認每個任務的具體要求。
2. **審查截圖和操作紀錄**：查看操作順序是否與指令一致。截圖如果與任務答案存在矛盾，優先選擇截圖。
3. **分析結果和任務答案**：考慮結果回應是否滿足所有任務要求。如果答案是"Maximum iteration limit exceeded, no answer found"，認定為失敗並分析可能的錯誤原因。
4. **提供改善建議**：根據分析，提出能夠增加成功率的規則。

# Output Format

你必須嚴格按照以下JSON格式輸出，不要添加任何額外的文本、說明或標記：

{
  "thought": "你的詳細推理。簡要總結截圖、結果回應和操作記錄中的證據，並解釋你的分析。",
  "answer": "'SUCCESS' 或 'NOT SUCCESS'",
  "rules": [
    {
      "rule_id": "serial number",
      "type": "Success Process | Special Phenomena/Mechanism | Error discover and how fix",
      "rule": "a rule of the game you discovered"
    }
  ]
}

# Notes

- 每個 Web 任務指令的所有部分是否如規定完成是需要特別關注的點。
- 不要與真實的網頁互動，也不要假設未提及的內容。
- 請確保返回的是有效的JSON格式，不包含任何其他文本。
"""
USER_PROMPT = """TASK: <task>
Result Response: <answer>
Operation Trace:
<assistant_process>
<num> screenshots at the end: """


def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def extract_assistant_process(messages):
    assistant_responses = []
    step = 1
    for msg in messages:
        if msg["role"] == "assistant" and "content" in msg:
            content = msg["content"]
            if isinstance(content, str) and ("Thought:" in content and "Action:" in content):
                thought = content.split("Action:")[0].replace("Thought:", "").strip()
                action = content.split("Action:")[1].strip()
                assistant_responses.append(f"Step {step}:\nThought: {thought}\nAction: {action}")
                step += 1
    return "\n\n".join(assistant_responses)


def auto_eval_by_gpt4v(process_dir, llm, img_num):
    print(f'--------------------- {process_dir} ---------------------')
    res_files = sorted(os.listdir(process_dir))
    with open(os.path.join(process_dir, 'interact_messages.json'), encoding='utf-8') as fr:
        it_messages = json.load(fr)

    if len(it_messages) == 1:
        print('Not find answer for ' + process_dir + ' only system messages')
        print()
        return {
            'result': 'NOT SUCCESS',
            'reason': 'Only system messages found, no assistant response',
            'task_question': 'Task information not found',
            'answer': '',
            'rules': []  # 添加空的rules字段
        }

    task_info = it_messages[1]["content"]
    if type(task_info) == list:
        task_info = task_info[0]["text"]
    assert 'Now given a task' in task_info
    pattern = r"Now given a task:(.+?)Please interact with"
    matches = re.search(pattern, task_info)
    task_content = matches.group(1).strip()
    # 提取任務問題（移除網址等額外資訊）
    pattern_question = r"(.+?)(?:\s*Please interact with|$)"
    match_question = re.search(pattern_question, task_content)
    task_question = match_question.group(1).strip() if match_question else task_content.strip()

    ans_info = it_messages[-1]["content"]
    if 'Action: ANSWER' not in ans_info:
        answer_content = "Maximum iteration limit exceeded, no answer found"
    else:
        pattern_ans = r"ANSWER[:; ]+\[?(.[^\]]*)\]?"
        matches_ans = re.search(pattern_ans, ans_info)
        answer_content = matches_ans.group(1).strip()

    # Extract assistant's thought process
    assistant_process = extract_assistant_process(it_messages)

    whole_content_img = []
    pattern_png = r'screenshot(\d+)\.png'
    matches = [(filename, int(re.search(pattern_png, filename).group(1))) for filename in res_files if re.search(pattern_png, filename)]
    matches.sort(key=lambda x: x[1])
    end_files = matches[-img_num:]
    for png_file in end_files:
        b64_img = encode_image(os.path.join(process_dir, png_file[0]))
        whole_content_img.append(
            {
                'type': 'image_url',
                'image_url': {"url": f"data:image/png;base64,{b64_img}"}
            }
        )

    user_prompt_tmp = USER_PROMPT.replace('<task>', task_content)
    user_prompt_tmp = user_prompt_tmp.replace('<answer>', answer_content)
    user_prompt_tmp = user_prompt_tmp.replace('<assistant_process>', assistant_process)
    user_prompt_tmp = user_prompt_tmp.replace('<num>', str(img_num))
    messages = [
        {'role': 'system', 'content': SYSTEM_PROMPT},
        {
            'role': 'user',
            'content': [
                {'type': 'text', 'text': user_prompt_tmp}
            ]
            + whole_content_img
            + [{'type': 'text', 'text': "Your verdict:\n"}]
        }
    ]
    while True:
        policyError = False
        try:
            print('Calling gpt4v API to get the auto evaluation......')
            # 添加response_format參數來強制輸出JSON格式
            response = llm.invoke(
                messages,
                response_format={"type": "json_object"}
            )
            token_usage = response.response_metadata.get('token_usage', {})
            prompt_tokens = token_usage.get('prompt_tokens', 0)
            completion_tokens = token_usage.get('completion_tokens', 0)

            print('Prompt Tokens:', prompt_tokens, ';',
                  'Completion Tokens:', completion_tokens)
            print('Cost:', prompt_tokens/1000 * 0.01
                  + completion_tokens / 1000 * 0.03)

            print('API call complete...')
            break
        except Exception as e:
            print(e)
            if type(e).__name__ == 'RateLimitError':
                time.sleep(10)
            elif type(e).__name__ == 'APIError':
                time.sleep(15)
            elif type(e).__name__ == 'InvalidRequestError':
                # 如果是格式問題，嘗試不使用response_format參數
                try:
                    print('Retrying without response_format parameter...')
                    response = llm.invoke(messages)
                    break
                except:
                    exit(0)
            elif "ResponsibleAIPolicyViolation" in str(e) and "content_filter" in str(e):
                print("Content ResponsibleAIPolicyViolation triggered. Breaking out of the loop.")
                policyError = True
                break
            else:
                time.sleep(10)
                
    # 解析回傳的JSON格式結果
    if policyError:
        return {
            'result': 'NOT SUCCESS',
            'reason': 'Content policy violation',
            'task_question': task_question,
            'answer': answer_content,
            'rules': []
        }
    
    gpt_4v_res = response.content
    print_message = messages[1]
    for idx in range(len(print_message['content'])):
        if print_message['content'][idx]['type'] == 'image_url':
            print_message['content'][idx]['image_url'] = {"url": "data:image/png;base64, b64_img"}

    print(print_message)
    print(gpt_4v_res)
    
    # 解析JSON回應 - 改進的解析邏輯
    try:
        # 如果回應是字符串，先嘗試直接解析
        if isinstance(gpt_4v_res, str):
            # 嘗試從可能的markdown格式中提取JSON
            json_match = re.search(r'```json\s*(.*?)\s*```', gpt_4v_res, re.DOTALL)
            if json_match:
                result_json = json.loads(json_match.group(1))
            else:
                # 嘗試直接解析整個字符串
                result_json = json.loads(gpt_4v_res)
        else:
            # 如果已經是字典類型，直接使用
            result_json = gpt_4v_res
        
        # 從JSON中提取結果
        thought = result_json.get('thought', '')
        answer = result_json.get('answer', '')
        rules = result_json.get('rules', [])
        
        # 確定成功或失敗
        result = 'NOT SUCCESS' if '不成功' in answer or '未成功' in answer else 'SUCCESS'
        
        return {
            'result': result,
            'reason': thought,  # 使用thought作為理由
            'task_question': task_question,
            'answer': answer_content,
            'rules': rules
        }
    except Exception as e:
        print(f"Error parsing JSON response: {e}")
        # 回退到原本的解析方式
        if '成功' in gpt_4v_res:
            auto_eval_res = 'SUCCESS'
        else:
            auto_eval_res = 'NOT SUCCESS'
        
        return {
            'result': auto_eval_res,
            'reason': gpt_4v_res,
            'task_question': task_question,
            'answer': answer_content,
            'rules': []  # 解析失敗時提供空的rules
        }

def save_evaluation_results(process_dir, results_by_website):
    """
    將評估結果儲存為 Excel 檔案，包含詳細結果和準確率統計
    Args:
        process_dir: 儲存結果的目錄路徑
        results_by_website: 包含每個網站評估結果的字典列表
    Returns:
        excel_filename: 儲存的 Excel 檔案路徑
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = os.path.join(process_dir, f'evaluation_results_{timestamp}.xlsx')
    
    wb = Workbook()
    # 設定詳細結果工作表
    ws_details = wb.active
    ws_details.title = "Detailed Results"
    
    # 設定欄位
    headers = ['Website', 'Task_ID', 'Task_Question', 'Result', 'Answer', 'Reason']
    for col, header in enumerate(headers, 1):
        ws_details.cell(row=1, column=col, value=header)
        # 設定欄寬
        ws_details.column_dimensions[get_column_letter(col)].width = 20 if col != 3 else 50
    
    # 寫入詳細結果
    for row, result in enumerate(results_by_website, 2):
        for col, header in enumerate(headers, 1):
            ws_details.cell(row=row, column=col, value=result[header])
    
    # 建立準確率統計工作表
    ws_accuracy = wb.create_sheet("Accuracy Statistics")
    ws_accuracy.append(['Website', 'Total Tasks', 'Successful Tasks', 'No Answer Tasks', 'Wrong Answer Tasks', 'Accuracy'])
    
    # 計算每個網站的準確率和各類任務數量
    website_stats = {}
    for result in results_by_website:
        website = result['Website']
        if website not in website_stats:
            website_stats[website] = {'total': 0, 'success': 0, 'no_answer': 0, 'wrong_answer': 0}
        website_stats[website]['total'] += 1
        if result['Result'] == 'SUCCESS':
            website_stats[website]['success'] += 1
        elif result['Result'] == 'NOT SUCCESS' and result['Reason'] == 'No final answer found in the conversation':
            website_stats[website]['no_answer'] += 1
    
    # 計算Wrong Answer數量並寫入準確率統計
    for website, stats in website_stats.items():
        # 計算Wrong Answer = Total - Success - No Answer
        stats['wrong_answer'] = stats['total'] - stats['success'] - stats['no_answer']
        accuracy = stats['success'] / stats['total'] if stats['total'] > 0 else 0
        ws_accuracy.append([
            website,
            stats['total'],
            stats['success'],
            stats['no_answer'],
            stats['wrong_answer'],
            f"{accuracy:.2%}"
        ])
    
    # 設定欄寬
    for col in range(1, 7):  # 更新欄數為6
        ws_accuracy.column_dimensions[get_column_letter(col)].width = 15
    
    # 新增一個工作表來儲存規則
    ws_rules = wb.create_sheet("Rules")
    ws_rules.append(['Website', 'Task_ID', 'Rule_ID', 'Type', 'Rule'])
    
    # 寫入規則
    rule_row = 2
    for result in results_by_website:
        website = result['Website']
        task_id = result['Task_ID']
        for rule in result.get('rules', []):
            ws_rules.append([
                website,
                task_id,
                rule.get('rule_id', ''),
                rule.get('type', ''),
                rule.get('rule', '')
            ])
            rule_row += 1
    
    # 設定規則工作表欄寬
    for col in range(1, 6):
        ws_rules.column_dimensions[get_column_letter(col)].width = 25 if col == 5 else 15
    
    wb.save(excel_filename)
    print(f"\nEvaluation results have been saved to: {excel_filename}")
    return excel_filename

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--process_dir', type=str, default='results')
    parser.add_argument('--lesson_dir', type=str, default='results')
    parser.add_argument("--api_key", default="key", type=str, help="YOUR_OPENAI_API_KEY")
    parser.add_argument("--api_model", default="gpt-4-vision-preview", type=str, help="api model name")
    parser.add_argument("--max_attached_imgs", type=int, default=1)
    parser.add_argument("--azure_endpoint", type=str, default="")
    parser.add_argument("--api_version", type=str, default="")
    parser.add_argument("--temperature", type=float, default=0.0, help="Temperature for the LLM response")
    parser.add_argument("--seed", type=int, default=42, help="Seed for random number generation")
    parser.add_argument("--llm", type=str, default="openai", choices=["openai", "azure", "openrouter", "gemini"])

    args = parser.parse_args()

    if args.llm == "openai":
        llm = ChatOpenAI(
            api_key=args.api_key,
            model=args.api_model,
            temperature=args.temperature
        )
    elif args.llm == "azure":
        llm = AzureChatOpenAI(
            api_key=args.api_key,
            model=args.api_model,
            api_version=args.api_version,
            temperature=args.temperature,
            azure_endpoint=args.azure_endpoint
        )
    elif args.llm == "openrouter":
        llm = ChatOpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key=args.api_key,
            model=args.api_model,
            temperature=args.temperature
        )
    elif args.llm == "gemini":
        #genai.configure(api_key=args.api_key)
        from langchain_google_genai import ChatGoogleGenerativeAI
        llm = ChatGoogleGenerativeAI(
            model=args.api_model,
            api_key=args.api_key,
            temperature=args.temperature,
            convert_system_message_to_human=True
        )

    all_results = []  # 儲存所有網站的評估結果
    
    webs = ['Allrecipes', 'Amazon', 'Apple', 'ArXiv', 'BBC News', 'Booking', 'Cambridge Dictionary',
            'Coursera', 'ESPN', 'GitHub', 'Google Flights', 'Google Map', 'Google Search', 'Huggingface', 'Wolfram Alpha']

    for web in webs:
        web_task_res = []  # list of dictionaries containing task id and response
        for idx in range(0, 46):
            file_dir = os.path.join(args.process_dir, 'task'+web+'--'+str(idx))
            if os.path.exists(file_dir):
                eval_result = auto_eval_by_gpt4v(file_dir, llm, args.max_attached_imgs)
                result_dict = {
                    'Website': web,
                    'Task_ID': idx,
                    'Task_Question': eval_result['task_question'],
                    'Result': eval_result['result'],
                    'Answer': eval_result['answer'],
                    'Reason': eval_result['reason'],
                    'rules': eval_result.get('rules', [])  # 新增rules欄位
                }
                web_task_res.append(result_dict)
                all_results.append(result_dict)
                
                # 同時打印到控制台
                print(f"task{web}--{idx}:")
                print(f"Question: {eval_result['task_question']}")
                print(f"Answer: {eval_result['answer']}")
                print(f"Result: {eval_result['result']}")
                print(f"Reason: {eval_result['reason']}")
                
                # 打印規則
                if eval_result.get('rules'):
                    print("Rules:")
                    for rule in eval_result['rules']:
                        print(f"  - [{rule.get('rule_id', '')}] {rule.get('type', '')}: {rule.get('rule', '')}")
                print("\n")
        
        if web_task_res:
            total_tasks = len(web_task_res)
            successful_tasks = sum(1 for res in web_task_res if res['Result'] == 'SUCCESS')
            accuracy = successful_tasks / total_tasks if total_tasks > 0 else 0
            print(f'{web} Accuracy: {accuracy:.2%}')

    # 儲存所有評估結果
    save_evaluation_results(args.process_dir, all_results)

if __name__ == '__main__':
    main()
