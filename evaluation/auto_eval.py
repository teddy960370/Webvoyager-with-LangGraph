import argparse
import os
import json
import time
import re
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from langchain_openai import AzureOpenAI,OpenAI,AzureChatOpenAI,ChatOpenAI

SYSTEM_PROMPT = """As an evaluator, you will be presented with four primary components to assist you in your role:

1. **Web Task Instruction**: A clear and specific directive provided in natural language, detailing the online activity to be carried out. These requirements may include conducting searches, verifying information, comparing prices, checking availability, or any other action relevant to the specified web service (such as Amazon, Apple, ArXiv, BBC News, Booking, etc.).

2. **Operation Trace**: A record of the assistant's thought process and the specific actions taken step-by-step during the task execution. 

3. **Result Screenshots**: A visual representation of the screen showing the result or intermediate state of performing a web task. It serves as visual proof of the actions taken in response to the instruction.

4. **Result Response**: A textual response obtained after the execution of the web task. It provides text-based results generated in response to the given instruction.

---

**Important Guidelines**:

- **DO NOT** interact with web pages or perform actions such as booking flights or conducting searches on real websites.
- **DO NOT** make assumptions based on information not presented in the screenshot, operation trace, or result response when comparing it to the instructions.
- Your primary responsibility is to conduct a thorough assessment of the web task instruction against the outcome depicted in the screenshots, the response text, and the operation trace, evaluating whether the actions taken align with the given instructions.
- **Note that the instruction may involve multiple tasks**. For example, locating the garage and summarizing the reviews. Failing to complete either task (e.g., not providing a summary) should be considered unsuccessful.
- Be aware of potential **discrepancies** between the provided "Result Response" and the "Result Screenshot":
  1. If the **Result Response contradicts the Screenshot**, the **content of the Screenshot prevails**.
  2. If the **Result Response contains content not mentioned in the Screenshot**, choose to **believe the content** of the **Result Response** unless otherwise specified.
- The **Operation Trace** should be reviewed to confirm that the sequence of actions aligns with the instructions and supports the provided results.

You should explicitly explain the reasoning behind your final evaluation before providing your verdict on whether the task has been successfully accomplished.

---

**Your reply MUST be in valid JSON format with the following structure**:

{
  "thought": "Your detailed reasoning. Briefly summarize the evidence from the Screenshot, Result Response, and Operation Trace, and explain your analysis.",
  "answer": "SUCCESS or NOT SUCCESS"
  "steps": [
    { 
        "Step": 1,
        "action": "The action taken in this step from operation Trace."
    },
    {   
        "Step": 2,
        "action": "The action taken in this step from operation Trace."
    },


  ]
}

---

# Notes

- Ensure each component (instruction, screenshots, result response, and operation trace) is considered in your reasoning before reaching a conclusion.
- Pay special attention to whether **all parts** of the Web Task Instruction have been completed as specified.
- You must only respond with valid JSON in the exact format specified above."""
USER_PROMPT = """TASK: <task>
Result Response: <answer>
Operation Trace:
<assistant_process>
<num> screenshots at the end: """


def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def extract_assistant_process(messages):
    assistant_responses = []
    step = 1
    for msg in messages:
        if msg["role"] == "assistant" and "content" in msg:
            content = msg["content"]
            if isinstance(content, str) and ("Thought:" in content and "Action:" in content):
                thought = content.split("Action:")[0].replace("Thought:", "").strip()
                action = content.split("Action:")[1].strip()
                assistant_responses.append(f"Step {step}:\nThought: {thought}\nAction: {action}")
                step += 1
    return "\n\n".join(assistant_responses) , step


def auto_eval_by_gpt4v(process_dir, llm, img_num):
    print(f'--------------------- {process_dir} ---------------------')
    res_files = sorted(os.listdir(process_dir))
    with open(os.path.join(process_dir, 'interact_messages.json'), encoding='utf-8') as fr:
        it_messages = json.load(fr)

    if len(it_messages) <= 1:
        print('Not find answer for ' + process_dir + ' only system messages')
        print()
        return {
            'result': 'NOT SUCCESS',
            'reason': 'Only system messages found, no assistant response',
            'task_question': 'Task information not found',
            'answer': '' , # 新增空的 answer 欄位
            'step_count': 0,
            'steps': []
        }

    task_info = it_messages[1]["content"]
    if type(task_info) == list:
        task_info = task_info[0]["text"]
    assert 'Now given a task' in task_info
    pattern = r"Now given a task:(.+?)Please interact with"
    matches = re.search(pattern, task_info)
    task_content = matches.group(1).strip()
    # 提取任務問題（移除網址等額外資訊）
    pattern_question = r"(.+?)(?:\s*Please interact with|$)"
    match_question = re.search(pattern_question, task_content)
    task_question = match_question.group(1).strip() if match_question else task_content.strip()

    ans_info = it_messages[-1]["content"]
    if 'Action: ANSWER' not in ans_info:
        print('Not find answer for ' + process_dir)
        print()
        return {
            'result': 'NOT SUCCESS',
            'reason': 'No final answer found in the conversation',
            'task_question': task_question,
            'answer': '',  # 新增空的 answer 欄位
            'step_count' : img_num,
            'steps': []
        }
    pattern_ans = r"ANSWER[:; ]+\[?(.[^\]]*)\]?"
    matches_ans = re.search(pattern_ans, ans_info)
    answer_content = matches_ans.group(1).strip()

    # Extract assistant's thought process
    assistant_process , step_count = extract_assistant_process(it_messages)

    # max_screenshot_id = max([int(f[10:].split('.png')[0]) for f in os.listdir(process_dir) if '.png' in f])
    # final_screenshot = f'screenshot{max_screenshot_id}.png'
    # b64_img = encode_image(os.path.join(process_dir, final_screenshot))
    whole_content_img = []
    pattern_png = r'screenshot(\d+)\.png'
    matches = [(filename, int(re.search(pattern_png, filename).group(1))) for filename in res_files if re.search(pattern_png, filename)]
    
    try :
        assert len(matches)  == step_count
    except AssertionError as e:
        print(f"Mismatch between number of screenshots and steps in the assistant process: {e}")


    matches.sort(key=lambda x: x[1])
    end_files = matches[-img_num:]
    for png_file in end_files:
        b64_img = encode_image(os.path.join(process_dir, png_file[0]))
        whole_content_img.append(
            {
                'type': 'image_url',
                'image_url': {"url": f"data:image/png;base64,{b64_img}"}
            }
        )

    user_prompt_tmp = USER_PROMPT.replace('<task>', task_content)
    user_prompt_tmp = user_prompt_tmp.replace('<answer>', answer_content)
    user_prompt_tmp = user_prompt_tmp.replace('<assistant_process>', assistant_process)
    user_prompt_tmp = user_prompt_tmp.replace('<num>', str(img_num))
    messages = [
        {'role': 'system', 'content': SYSTEM_PROMPT},
        {
            'role': 'user',
            'content': [
                {'type': 'text', 'text': user_prompt_tmp}
            ]
            + whole_content_img
            + [{'type': 'text', 'text': "Your verdict:\n"}]
        }
    ]
    while True:
        policyError = False
        try:
            print('Calling gpt4v API to get the auto evaluation......')
            # 添加response_format參數強制JSON輸出
            response = llm.invoke(
                messages,
                response_format={"type": "json_object"}
            )
            token_usage = response.response_metadata.get('token_usage', {})
            prompt_tokens = token_usage.get('prompt_tokens', 0)
            completion_tokens = token_usage.get('completion_tokens', 0)

            print('Prompt Tokens:', prompt_tokens, ';',
                  'Completion Tokens:', completion_tokens)
            print('Cost:', prompt_tokens/1000 * 0.01
                  + completion_tokens / 1000 * 0.03)

            print('API call complete...')
            break
        except Exception as e:
            print(e)
            if type(e).__name__ == 'RateLimitError':
                time.sleep(10)
            elif type(e).__name__ == 'APIError':
                time.sleep(15)
            elif type(e).__name__ == 'InvalidRequestError':
                # 嘗試不使用response_format參數再試一次
                try:
                    print('Retrying without response_format parameter...')
                    response = llm.invoke(messages)
                    break
                except:
                    exit(0)
            elif "ResponsibleAIPolicyViolation" in str(e) and "content_filter" in str(e):
                print("Content ResponsibleAIPolicyViolation triggered. Breaking out of the loop.")
                policyError = True
                break
            else:
                time.sleep(10)
                
    if policyError:
        return {
            'result': 'NOT SUCCESS',
            'reason': 'Content policy violation',
            'task_question': task_question,
            'answer': answer_content,
            'step_count': step_count,
            'steps': []
        }
    
    gpt_4v_res = response.content
    print_message = messages[1]
    for idx in range(len(print_message['content'])):
        if print_message['content'][idx]['type'] == 'image_url':
            print_message['content'][idx]['image_url'] = {"url": "data:image/png;base64, b64_img"}

    #print(print_message)
    #print(gpt_4v_res)

    # 使用新的JSON解析邏輯
    try:
        # 嘗試解析JSON回應
        if isinstance(gpt_4v_res, dict):
            # 如果已經是字典，直接使用
            result_json = gpt_4v_res
        else:
            # 如果是字符串，嘗試提取JSON部分
            json_match = re.search(r'```json\s*(.*?)\s*```', gpt_4v_res, re.DOTALL)
            if json_match:
                result_json = json.loads(json_match.group(1))
            else:
                # 嘗試直接解析整個字符串
                result_json = json.loads(gpt_4v_res)
        
        # 從JSON中提取結果
        thought = result_json.get('thought', '')
        answer = result_json.get('answer', '')
        steps = result_json.get('steps', [])
        
        # 確定成功或失敗
        auto_eval_res = 'SUCCESS' if answer.upper() == 'SUCCESS' else 'NOT SUCCESS'
        
        print('Auto_eval_res:', auto_eval_res)
        print('Evaluation Reason:', thought)
        
        return {
            'result': auto_eval_res,
            'reason': thought,
            'task_question': task_question,
            'answer': answer_content,
            'step_count': step_count,
            'steps': steps
        }
    except Exception as e:
        print(f"Error parsing JSON response: {e}")
        return

def save_evaluation_results(process_dir, results_by_website , max_steps):
    """
    將評估結果儲存為 Excel 檔案，包含詳細結果和準確率統計
    Args:
        process_dir: 儲存結果的目錄路徑
        results_by_website: 包含每個網站評估結果的字典列表
    Returns:
        excel_filename: 儲存的 Excel 檔案路徑
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = os.path.join(process_dir, f'evaluation_results_{timestamp}.xlsx')
    
    wb = Workbook()
    # 設定詳細結果工作表
    ws_details = wb.active
    ws_details.title = "Detailed Results"
    
    # 設定欄位
    headers = ['Website', 'Task_ID', 'Task_Question', 'Result', 'Answer', 'Reason' , 'Steps']
    for col, header in enumerate(headers, 1):
        ws_details.cell(row=1, column=col, value=header)
        # 設定欄寬
        ws_details.column_dimensions[get_column_letter(col)].width = 20 if col != 3 else 50
    
    # 寫入詳細結果
    for row, result in enumerate(results_by_website, 2):
        for col, header in enumerate(headers, 1):
            if header == 'Steps':
                # 將步驟列表轉換為字符串
                steps_str = '\n'.join([f"Step {step['Step']}: {step['action']}" for step in result['Steps']])
                ws_details.cell(row=row, column=col, value=steps_str)
            else:
                ws_details.cell(row=row, column=col, value=result[header])
    
    # 建立準確率統計工作表
    ws_accuracy = wb.create_sheet("Accuracy Statistics")
    ws_accuracy.append(['Website', 'Total Tasks', 'Successful Tasks', 'No Answer Tasks', 'Wrong Answer Tasks', 'Step(Avg.)', 'Accuracy'])
    
    # 計算每個網站的準確率和各類任務數量
    website_stats = {}
    for result in results_by_website:
        website = result['Website']
        if website not in website_stats:
            website_stats[website] = {'total': 0, 'success': 0, 'no_answer': 0, 'wrong_answer': 0 , 'total_steps': 0}
        website_stats[website]['total'] += 1
        website_stats[website]['total_steps'] += len(result['Steps'])  # 累計步驟數
        if result['Result'] == 'no_answer':
            website_stats[website]['total_steps'] += max_steps  # no answer is reach max steps
        if result['Result'] == 'SUCCESS':
            website_stats[website]['success'] += 1
        elif result['Result'] == 'NOT SUCCESS' and result['Reason'] == 'No final answer found in the conversation':
            website_stats[website]['no_answer'] += 1
    
    # 計算Wrong Answer數量並寫入準確率統計
    for website, stats in website_stats.items():
        # 計算Wrong Answer = Total - Success - No Answer
        stats['wrong_answer'] = stats['total'] - stats['success'] - stats['no_answer']
        accuracy = stats['success'] / stats['total'] if stats['total'] > 0 else 0
        # 計算平均步驟數
        avg_steps = stats['total_steps'] / stats['total'] if stats['total'] > 0 else 0
        ws_accuracy.append([
            website,
            stats['total'],
            stats['success'],
            stats['no_answer'],
            stats['wrong_answer'],
            f"{avg_steps:.2f}",
            f"{accuracy:.2%}"
        ])
    
    # 設定欄寬
    for col in range(1, 7):  # 更新欄數為6
        ws_accuracy.column_dimensions[get_column_letter(col)].width = 15
    
    wb.save(excel_filename)
    print(f"\nEvaluation results have been saved to: {excel_filename}")
    return excel_filename

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--process_dir', type=str, default='results')
    parser.add_argument('--lesson_dir', type=str, default='results')
    parser.add_argument("--api_key", default="key", type=str, help="YOUR_OPENAI_API_KEY")
    parser.add_argument("--api_model", default="gpt-4-vision-preview", type=str, help="api model name")
    parser.add_argument("--max_attached_imgs", type=int, default=1)
    parser.add_argument("--azure_endpoint", type=str, default="")
    parser.add_argument("--api_version", type=str, default="")
    parser.add_argument("--temperature", type=float, default=0.0, help="Temperature for the LLM response")
    parser.add_argument("--seed", type=int, default=42, help="Seed for random number generation")
    parser.add_argument("--llm", type=str, default="openai", choices=["openai", "azure", "openrouter", "gemini"])
    parser.add_argument('--max_iter', type=int, default=15)

    args = parser.parse_args()

    if args.llm == "openai":
        llm = ChatOpenAI(
            api_key=args.api_key,
            model=args.api_model,
            temperature=args.temperature
        )
    elif args.llm == "azure":
        llm = AzureChatOpenAI(
            api_key=args.api_key,
            model=args.api_model,
            api_version=args.api_version,
            temperature=args.temperature,
            azure_endpoint=args.azure_endpoint
        )
    elif args.llm == "openrouter":
        llm = ChatOpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key=args.api_key,
            model=args.api_model,
            temperature=args.temperature
        )
    elif args.llm == "gemini":
        #genai.configure(api_key=args.api_key)
        from langchain_google_genai import ChatGoogleGenerativeAI
        llm = ChatGoogleGenerativeAI(
            model=args.api_model,
            api_key=args.api_key,
            temperature=args.temperature,
            convert_system_message_to_human=True
        )

    all_results = []  # 儲存所有網站的評估結果
    
    webs = ['Allrecipes', 'Amazon', 'Apple', 'ArXiv', 'BBC News', 'Booking', 'Cambridge Dictionary',
            'Coursera', 'ESPN', 'GitHub', 'Google Flights', 'Google Map', 'Google Search', 'Huggingface', 'Wolfram Alpha']

    for web in webs:
        web_task_res = []  # list of dictionaries containing task id and response
        for idx in range(0, 46):
            file_dir = os.path.join(args.process_dir, 'task'+web+'--'+str(idx))
            if os.path.exists(file_dir):
                eval_result = auto_eval_by_gpt4v(file_dir, llm, args.max_attached_imgs)  # 修正參數從client到llm
                result_dict = {
                    'Website': web,
                    'Task_ID': idx,
                    'Task_Question': eval_result['task_question'],
                    'Result': eval_result['result'],
                    'Answer': eval_result['answer'],  # 新增 answer 欄位
                    'Reason': eval_result['reason'],
                    'Steps': eval_result['steps']
                }
                web_task_res.append(result_dict)
                all_results.append(result_dict)
                
                # 同時打印到控制台
                print(f"task{web}--{idx}:")
                print(f"Question: {eval_result['task_question']}")
                print(f"Answer: {eval_result['answer']}")
                print(f"Result: {eval_result['result']}")
                print(f"Reason: {eval_result['reason']}\n")


                # save results as json file
                knowledge_dir = f"./data/{web}"
                if not os.path.exists(knowledge_dir):
                    os.makedirs(knowledge_dir)
                fileName = f"task{web}--{idx}_{eval_result['result']}_eval.json"
                with open(os.path.join(knowledge_dir, fileName), 'w', encoding='utf-8') as f:
                    json.dump(result_dict, f, ensure_ascii=False, indent=4)

        if web_task_res:
            total_tasks = len(web_task_res)
            successful_tasks = sum(1 for res in web_task_res if res['Result'] == 'SUCCESS')
            accuracy = successful_tasks / total_tasks if total_tasks > 0 else 0
            print(f'{web} Accuracy: {accuracy:.2%}')

    # 儲存所有評估結果

    save_evaluation_results(args.process_dir, all_results , args.max_iter)

    


if __name__ == '__main__':
    main()
